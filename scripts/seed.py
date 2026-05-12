#!/usr/bin/env python3
"""Seed Supabase with data from the nSFR Excel file."""
import sys
import json
import urllib.request
import urllib.error
import ssl
import openpyxl
import collections
from datetime import datetime

# Bypass SSL verification for local seeding script
ssl_ctx = ssl.create_default_context()
ssl_ctx.check_hostname = False
ssl_ctx.verify_mode = ssl.CERT_NONE

SUPABASE_URL = "https://gyttyxzqymutewfmmgis.supabase.co"
SERVICE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imd5dHR5eHpxeW11dGV3Zm1tZ2lzIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NzUxOTQwMiwiZXhwIjoyMDkzMDk1NDAyfQ._XtdxkBVOU-SsWc2IiAyaf7VPYpV4P1W5wCGQes8c44"

HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal",
}

def post(table, rows):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    data = json.dumps(rows).encode()
    req = urllib.request.Request(url, data=data, headers=HEADERS, method="POST")
    try:
        with urllib.request.urlopen(req, context=ssl_ctx) as r:
            return r.status
    except urllib.error.HTTPError as e:
        print(f"  ERROR {e.code}: {e.read().decode()[:200]}")
        return e.code

def safe_int(v):
    if v is None or str(v).strip().lower() in ("null", "", "none"):
        return None
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return None

def safe_float(v):
    if v is None or str(v).strip().lower() in ("null", "", "none"):
        return 0.0
    try:
        return float(str(v))
    except (ValueError, TypeError):
        return 0.0

def safe_str(v):
    if v is None or str(v).strip().lower() in ("null",):
        return None
    return str(v).strip() or None

def serial_to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, str):
        try:
            return datetime.strptime(v.split("T")[0], "%Y-%m-%d").date().isoformat()
        except Exception:
            return None
    return None

def serial_to_dt(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, str):
        try:
            return datetime.fromisoformat(v).isoformat()
        except Exception:
            return None
    return None

BATCH_ID = "seed_v2_may2026"
FILE = "/Users/john/Desktop/nSFR.xlsx"

# ---- CLEAR EXISTING DATA ----
def delete_all(table):
    url = f"{SUPABASE_URL}/rest/v1/{table}?id=gte.0"
    req = urllib.request.Request(url, headers={**HEADERS, "Prefer": "return=minimal"}, method="DELETE")
    try:
        with urllib.request.urlopen(req, context=ssl_ctx) as r:
            print(f"  Cleared {table}: {r.status}")
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        # Fall back to delete-all via neq trick
        url2 = f"{SUPABASE_URL}/rest/v1/{table}?upload_batch_id=neq.____never____"
        req2 = urllib.request.Request(url2, headers={**HEADERS, "Prefer": "return=minimal"}, method="DELETE")
        try:
            with urllib.request.urlopen(req2, context=ssl_ctx) as r2:
                print(f"  Cleared {table} (fallback): {r2.status}")
        except Exception as e2:
            print(f"  Clear {table} failed: {e2}")

print("Clearing existing data...")
delete_all("refunds")
delete_all("product_failures")
delete_all("upload_batches")

print(f"Loading {FILE}...")
wb = openpyxl.load_workbook(FILE)

# ---- REFUNDS ----
ws = wb["Refunds"]
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
col = {h: i for i, h in enumerate(headers)}
data = rows[1:]

refund_rows = []
for r in data:
    week_val = r[col["week"]]
    try:
        week_int = int(week_val) if week_val is not None else None
    except (ValueError, TypeError):
        week_int = None

    refund_rows.append({
        "country": safe_str(r[col["Country"]]) or "Kenya",
        "store": safe_str(r[col["Store"]]) or "",
        "order_date": serial_to_date(r[col["Order Date"]]),
        "order_start_time": serial_to_dt(r[col["Order Start Time (local)"]]),
        "order_dropoff_time": serial_to_dt(r[col["Order Drop-off by Rider (local)"]]),
        "order_id": safe_int(r[col["Order Id"]]),
        "split_stacked": safe_str(r[col["Split / Stacked"]]),
        "picker_id": safe_int(r[col["Picker Id"]]),
        "picker_name": safe_str(r[col["Picker Name"]]),
        "rider_id": safe_int(r[col["Rider Id"]]),
        "sku": safe_str(r[col["SKU"]]),
        "product_name": safe_str(r[col["Product Name"]]),
        "category_l1": safe_str(r[col["Category L1"]]),
        "category_l2": safe_str(r[col["Category L2"]]),
        "supplier_name": safe_str(r[col["Supplier Name"]]),
        "event": safe_str(r[col["Event"]]),
        "ccr3": safe_str(r[col["CCR3"]]),
        "origin": safe_str(r[col["origin"]]),
        "refund": safe_float(r[col["Refund"]]),
        "compensation": safe_float(r[col["Compensation"]]),
        "refund_and_comp": safe_float(r[col["Refund & Comp"]]),
        "week": week_int,
        "upload_batch_id": BATCH_ID,
    })

print(f"Inserting {len(refund_rows)} refund rows...")
CHUNK = 200
for i in range(0, len(refund_rows), CHUNK):
    chunk = refund_rows[i:i+CHUNK]
    status = post("refunds", chunk)
    print(f"  Chunk {i//CHUNK + 1}/{(len(refund_rows)+CHUNK-1)//CHUNK}: {status}")

# ---- PRODUCT FAILURES ----
# Sheet is "PFR" in the new file format
pf_sheet_name = next((n for n in wb.sheetnames if "pfr" in n.lower() or "product failure" in n.lower() or "product_failure" in n.lower()), None)
if not pf_sheet_name:
    pf_sheet_name = wb.sheetnames[0] if len(wb.sheetnames) > 0 else "PFR"
ws2 = wb[pf_sheet_name]
print(f"Using PF sheet: {pf_sheet_name}")
rows2 = list(ws2.iter_rows(values_only=True))
headers2 = rows2[0]
col2 = {h: i for i, h in enumerate(headers2)}
data2 = rows2[1:]

pf_rows = []
for r in data2:
    week_val = r[col2["Week"]]
    try:
        week_int = int(week_val) if week_val is not None else None
    except (ValueError, TypeError):
        week_int = None

    pf_rows.append({
        "store": safe_str(r[col2["Store"]]) or "",
        "order_id": safe_int(r[col2["Order ID"]]),
        "order_date": serial_to_date(r[col2["order_placed_localtime_at_date"]]),
        "sku": safe_str(r[col2["SKU"]]),
        "sku_name": safe_str(r[col2["SKU Name"]]),
        "qty_ordered": safe_int(r[col2["Qty Ordered"]]),
        "qty_delivered": safe_int(r[col2["Qty Delivered"]]),
        "on_hand_qty_before": safe_int(r[col2["On Hand Qty Before Order"]]),
        "on_hand_qty_delta": safe_int(r[col2["On Hand Qty Delta"]]),
        "on_hand_qty_after": safe_int(r[col2["On Hand Qty After Order"]]),
        "reserved_qty_before": safe_int(r[col2["Reserved Qty Before Order"]]),
        "reserved_qty_delta": safe_int(r[col2["Reserved Qty Delta"]]),
        "sales_buffer": safe_int(r[col2["Sales Buffer"]]),
        "im_avail": safe_int(r[col2["IM Avail."]]),
        "im_avail_minus_qty_ord": safe_int(r[col2["IM Avail. - Qty Ord"]]),
        "pf_root_cause": safe_str(r[col2["PF Root Cause"]]),
        "week": week_int,
        "picker_name": safe_str(r[col2["Picker Name"]]),
        "upload_batch_id": BATCH_ID,
    })

print(f"Inserting {len(pf_rows)} PF rows...")
for i in range(0, len(pf_rows), CHUNK):
    chunk = pf_rows[i:i+CHUNK]
    status = post("product_failures", chunk)
    print(f"  Chunk {i//CHUNK + 1}/{(len(pf_rows)+CHUNK-1)//CHUNK}: {status}")

# ---- Batch record ----
all_stores = list(set(r["store"] for r in refund_rows + pf_rows if r.get("store")))
all_weeks = [r["week"] for r in refund_rows + pf_rows if r.get("week")]
min_w = min(all_weeks) if all_weeks else 1
max_w = max(all_weeks) if all_weeks else 17

batch = [{
    "id": BATCH_ID,
    "filename": "nSFR Data.xlsx",
    "uploaded_at": datetime.now().isoformat(),
    "refund_count": len(refund_rows),
    "pf_count": len(pf_rows),
    "stores": all_stores,
    "week_range": f"W{min_w}–W{max_w}",
}]
status = post("upload_batches", batch)
print(f"Batch record: {status}")
print("Done!")
